import openpyxl
from openpyxl.styles import Font

def crear_modelo_financiero():
    wb = openpyxl.Workbook()

    # --- 1. Hoja Ventas ---
    ws_ventas = wb.active
    ws_ventas.title = "Ventas"
    
    datos_ventas = [
        ["PRESUPUESTO DE VENTAS (PROYECCIÓN MENSUAL)"],
        [],
        ["Línea de Servicio", "Precio Unitario", "Cantidad", "Venta Total", "Costo Variable", "Margen Contribución"],
        ["Desarrollo de Software a la medida", 3500000, 2, 7000000, 300000, 3200000],
        ["Mantenimiento y Soporte (Mensual)", 300000, 5, 1500000, 50000, 250000],
        ["TOTALES", "", 7, 8500000, "", ""],
        [],
        ["COSTOS FIJOS OPERATIVOS"],
        ["Concepto", "Valor Mensual"],
        ["Servicios públicos (Energía e Internet)", 150000],
        ["Suscripciones base (Hosting, Repositorios)", 100000],
        ["Imprevistos / Papelería", 50000],
        ["TOTAL COSTOS FIJOS", 300000]
    ]
    
    for fila in datos_ventas:
        ws_ventas.append(fila)
        
    for r in [1, 3, 6, 8, 9, 13]:
        for cell in ws_ventas[r]:
            cell.font = Font(bold=True)

    # --- 2. Hoja Personal ---
    ws_personal = wb.create_sheet(title="Personal")
    datos_personal = [
        ["PRESUPUESTO DE PERSONAL"],
        [],
        ["Cargo", "Cantidad", "Salario Base", "Factor Prestacional (52%)", "Costo Individual", "Costo Consolidado"],
        ["Desarrollador Principal", 1, 2000000, 1040000, 3040000, 3040000],
        ["Asesor Comercial (Medio tiempo)", 1, 800000, 416000, 1216000, 1216000],
        ["TOTAL NÓMINA", 2, "", "", "", 4256000]
    ]
    
    for fila in datos_personal:
        ws_personal.append(fila)
        
    for r in [1, 3, 6]:
        for cell in ws_personal[r]:
            cell.font = Font(bold=True)

    # --- 3. Hoja Inversión ---
    ws_inv = wb.create_sheet(title="Inversión")
    datos_inv = [
        ["ACTIVOS FIJOS"],
        [],
        ["Descripción", "Cantidad", "Valor Unitario", "Valor Total"],
        ["Equipo de Cómputo", 1, 3200000, 3200000],
        ["Monitor Secundario y Perif.", 1, 600000, 600000],
        ["Mobiliario", 1, 700000, 700000],
        ["TOTAL ACTIVOS FIJOS", "", "", 4500000],
        [],
        ["CAPITAL INICIAL Y FUENTES"],
        [],
        ["Concepto", "Valor Requerido", "Tipo de Fuente", "Origen Específico"],
        ["Activos Fijos", 4500000, "Capital Propio", "Ahorros personales"],
        ["Capital de Trabajo", 5500000, "Externo (Subsidio)", "Fondo Emprender SENA"],
        ["CAPITAL INICIAL TOTAL", 10000000, "", ""]
    ]
    
    for fila in datos_inv:
        ws_inv.append(fila)
        
    for r in [1, 3, 7, 9, 11, 14]:
        for cell in ws_inv[r]:
            cell.font = Font(bold=True)

    # --- 4. Hoja Estados Financieros ---
    ws_fin = wb.create_sheet(title="Estados Financieros")
    datos_fin = [
        ["BALANCE INICIAL"],
        [],
        ["ACTIVO", "Valor", "PASIVO Y PATRIMONIO", "Valor"],
        ["Activo Corriente", "", "Pasivo", ""],
        ["Caja / Bancos", 5500000, "Obligaciones financieras", 0],
        ["", "", "Cuentas por pagar", 0],
        ["Activo No Corriente", "", "Total Pasivos", 0],
        ["Equipos de Computación", 3800000, "Patrimonio", ""],
        ["Muebles y Enseres", 700000, "Capital Social", 10000000],
        ["TOTAL ACTIVOS", 10000000, "TOTAL PASIVO + PATRIMONIO", 10000000],
        [],
        [],
        ["FLUJO DE CAJA (TRIMESTRAL)"],
        [],
        ["Concepto", "Mes 1", "Mes 2", "Mes 3"],
        ["INGRESOS", "", "", ""],
        ["Saldo Inicial en Caja", 5500000, 8594000, 11688000],
        ["Ingresos por Ventas", 8500000, 8500000, 9500000],
        ["Total Ingresos Disponibles (A)", 14000000, 17094000, 21188000],
        [],
        ["EGRESOS", "", "", ""],
        ["Costos Variables", 850000, 850000, 900000],
        ["Costos Fijos", 300000, 300000, 300000],
        ["Nómina y Prestaciones", 4256000, 4256000, 4256000],
        ["Total Egresos (B)", 5406000, 5406000, 5456000],
        [],
        ["SALDO FINAL EN CAJA (A - B)", 8594000, 11688000, 15732000]
    ]
    
    for fila in datos_fin:
        ws_fin.append(fila)
        
    for r in [1, 3, 4, 7, 10, 13, 15, 16, 19, 21, 25, 27]:
        for cell in ws_fin[r]:
            cell.font = Font(bold=True)

    # --- Ajustes Globales ---
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                # Aplicar formato de moneda a números mayores a 100 o ceros financieros
                if isinstance(cell.value, int) and (cell.value > 100 or cell.value == 0):
                    cell.number_format = '"$"#,##0'
        
        # Ajuste automático del ancho de las columnas
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

    # Guardar el documento final
    nombre_archivo = "GA10_240201529_AA3_EV01_Modelo_Financiero.xlsx"
    wb.save(nombre_archivo)
    print(f"¡Evidencia generada exitosamente como '{nombre_archivo}'!")

if __name__ == "__main__":
    crear_modelo_financiero()